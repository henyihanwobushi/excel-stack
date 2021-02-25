from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.copier import WorksheetCopy
from xls2xlsx import XLS2XLSX
from copy import copy, deepcopy
import ntpath
import xlrd
from gooey import Gooey, GooeyParser

@Gooey(
    advanced=True,
    required_cols=1,
    optional_cols=1,
    use_cmd_args=True,
    dump_build_config=True,
)
def main():
    parser = GooeyParser(description="Worksheet Stack")
    parser.add_argument('input-filenames', nargs='+', widget='MultiFileChooser')
    parser.add_argument('--output-dir', widget='DirChooser')
    args = parser.parse_args()
    stack(input_filenames=args.__dict__['input-filenames'], output_dir=args.output_dir)


def stack(input_filenames=[], output_dir=""):
    """  """
    output_dir = output_dir.rstrip('/')
    output_files = []
    output_file_formated = []
    for i in range(len(input_filenames)):
        wb = read_table(input_filenames[i])
        while len(output_files) < len(wb.sheetnames):
            w = Workbook()
            w.remove(w[w.sheetnames[0]])
            output_files.append(w)
            output_file_formated.append(False)
        for k in range(len(wb.sheetnames)):
            copy_sheet(wb, wb.sheetnames[k], output_files[k], get_filename(input_filenames[i]))
            if output_file_formated[k]:
                continue
            copy_alignments(wb, output_files[k])
            copy_protection(wb, output_files[k])
            copy_fills(wb, output_files[k])
            copy_fonts(wb, output_files[k])
            copy_borders(wb, output_files[k])
            output_file_formated[k] = True


    for j in range(len(output_files)):
        output_filename = f'{output_dir}/sheet_{j}.xlsx'
        output_files[j].save(filename = output_filename)


def read_table(filename):
    if len(filename) > 5 and filename[-5:] == '.xlsx':
        return load_workbook(filename)
    elif len(filename) > 4 and filename[-4:] == '.xls':
        return read_xls(filename)
    else:
        raise Exception("not supported file format")

def copy_alignments(work_book1, work_book2):
    work_book2._alignments = [copy(a) for a in work_book1._alignments]

def copy_protection(work_book1, work_book2):
    work_book2._protections = [copy(a) for a in work_book1._protections]

def copy_protection(work_book1, work_book2):
    work_book2._protections = [copy(a) for a in work_book1._protections]

def copy_borders(work_book1, work_book2):
    work_book2._borders = [copy(a) for a in work_book1._borders]

def copy_fills(work_book1, work_book2):
    work_book2._fills = [copy(a) for a in work_book1._fills]

def copy_fonts(work_book1, work_book2):
    work_book2._fonts = [copy(a) for a in work_book1._fonts]

def get_filename(fn):
    bn = ntpath.basename(fn)
    return bn[:bn.index('.')]


def copy_sheet(work_book1, sheet_name1, work_book2, sheet_name2):
    ws2 = work_book2.create_sheet(sheet_name2)
    ws1 = work_book1[sheet_name1]
    copy_cells(ws1, ws2)
    copy_dimensions(ws1, ws2)

    ws2.sheet_format = copy(ws1.sheet_format)
    ws2.sheet_properties = copy(ws1.sheet_properties)
    ws2.merged_cells = copy(ws1.merged_cells)
    ws2.page_margins = copy(ws1.page_margins)
    ws2.page_setup = copy(ws1.page_setup)
    ws2.print_options = copy(ws1.print_options)


def copy_cells(worksheet_1, worksheet_2):
    for (row, col), source_cell  in worksheet_1._cells.items():
        target_cell = worksheet_2.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        # target_cell.fill = copy(source_cell.fill)
        # target_cell.font = copy(source_cell.font)
        # target_cell.border = copy(source_cell.border)
        target_cell._style = copy(source_cell._style)
        target_cell._hyperlink = copy(source_cell.hyperlink)
        target_cell.comment = copy(source_cell.comment)


def copy_dimensions(worksheet_1, worksheet_2):
    for attr in ('row_dimensions', 'column_dimensions'):
        src = getattr(worksheet_1, attr)
        target = getattr(worksheet_2, attr)
        for key, dim in src.items():
            target[key] = copy(dim)
            target[key].worksheet = worksheet_2


def read_xls(filename):
    """ read xls to openyxl workbook """
    # first open using xlrd
    book = XLS2XLSX(filename)
    return book.to_xlsx()


if __name__ == '__main__':
    main()
